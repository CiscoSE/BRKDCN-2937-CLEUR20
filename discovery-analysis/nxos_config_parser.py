# Copyright (c) 2020 Cisco and/or its affiliates.
#
# This software is licensed to you under the terms of the Cisco Sample
# Code License, Version 1.1 (the "License"). You may obtain a copy of the
# License at
#
#                https://developer.cisco.com/docs/licenses
#
# All use of the material herein must be in accordance with the terms of
# the License. All rights not expressly granted by the License are
# reserved. Unless required by applicable law or agreed to separately in
# writing, software distributed under the License is distributed on an "AS
# IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
# or implied.

#-------------------------------------------------------------------------------
import re
import argparse
import ipaddress
import os
from pprint import pprint
from  openpyxl import Workbook
from openpyxl.worksheet.table import Table , TableStyleInfo
import detect_os_type_and_version as detect_os
#-------------------------------------------------------------------------------

def read_config_file(file_name):
    lines = []
    try:
        for line in open(file_name):
            lines.append(line)
        return lines
    except IOError:
        print("ERROR")
    except Exception:
        print("ERROR")

def grep(pattern, line):
    if re.search(r'%s'%pattern, line.rstrip(), flags=0):
        return True
    return False

def get_indent(line):
    return len(line) - len(line.lstrip())

def expand_ip_range(start_range, stop_range):
    """
    """
    subnets_list = []
    start_range_bytes = start_range.split(".")
    stop_range_bytes = stop_range.split(".")
    last_byte_dif = int(stop_range_bytes[3])- int(start_range_bytes[3])
    base_range = ".".join(start_range_bytes[0:3])
    for i in range(0, last_byte_dif+1):
        subnet = base_range + "." + str(int(start_range_bytes[3])+i) +"/32"
        subnets_list.append(subnet)
    return subnets_list

def split_expand_vlan_range(vlan_string):
    """
        Parse a string in the form of vlan 1,4,8,12-14
        And produce a list of the form [1,4,8,12,13,14]
    """
    raw_vlan_list = vlan_string.lstrip("vlan").lstrip().rstrip().split(",")
    vlan_id_list = []
    for vlan_id in raw_vlan_list:
        if re.search(r'[0-9]+-[0-9]+', vlan_id, flags=0):
            vlan_range = vlan_id.split("-")
            for i in range(int(vlan_range[0]), int(vlan_range[1]) + 1):
                vlan_id_list.append(str(i))
        else:
            vlan_id_list.append(str(vlan_id))
    return vlan_id_list

def print_to_csv(headers, output_dictionary, output_file):
    """
    """
    with open(output_file, 'w') as f:
        print("+--- Printing output---+")
        header_line = headers[0]
        for header in headers[1:]:
            header_line = header_line + ";" + header
        f.write(header_line+"\n")
        for key in output_dictionary.keys():
            new_line = output_dictionary[key][headers[0]]
            for header in headers[1:]:
                new_line = new_line + ";" + str(output_dictionary[key][header])
            f.write(new_line + "\n")

def create_worksheet(excel_workbook, headers, output_dictionary, sheet_name):
    """
    """
    sheet = excel_workbook.create_sheet(title=sheet_name)
    for i in range(0, len(headers)):
        sheet.cell(column=i+1, row=1, value=headers[i])
    row_id = 2
    for key in output_dictionary.keys():
        for i in range(0, len(headers)):
            if isinstance (output_dictionary[key][headers[i]],list):
                sheet.cell(column=i+1, row=row_id, value="{0}".format(",".join(output_dictionary[key][headers[i]])))
            else:
                sheet.cell(column=i+1, row=row_id, value="{0}".format(output_dictionary[key][headers[i]]))
        row_id = row_id + 1
    if len(headers) > 26:  ### Super Ugly !! Fix That 
        table_cells = "A1:A" + chr(64+len(headers)-26)
    else:
        table_cells = "A1:" + chr(64+len(headers)) + str(row_id)
    style = TableStyleInfo(name = "TableStyleMedium9" , showRowStripes="True" )
    table = Table(displayName = sheet_name , ref = table_cells)
    table.tableStyleInfo = style
    sheet.add_table(table)
    return excel_workbook

def parse_fex_port(config_files):
    """
    Parsing the following format

    interface Ethernet1/7
        description 10GE fabric to mnd-rc0002-ds173-101 - po101 - MND-RC0002-MND-RC0002-EBRR
        switchport
        switchport mode fex-fabric
        fex associate 101
        channel-group 101
        no shutdown

    interface Ethernet1/8
        description enc1-vcenet1-mgt.mnd.bss.local interface X1
        switchport
        switchport mode trunk
        switchport trunk allowed vlan 101,300,310,320,330,340,370,410,430,440,450,460,468,470,510,512,520,530,560,610,620,630,640,650,660,670,690,710,720,730,750,760,770,780,790,801-804,806-807,1014,1233,1710,1712,1720,3018
        spanning-tree port type edge trunk
        no shutdown
    interface Ethernet1/27
        description hnas2.mnd.office.intern interface 10GbE 4
        switchport
        switchport mode trunk
        switchport trunk allowed vlan 440,450,460,750
        spanning-tree port type edge
        spanning-tree bpdufilter enable
        channel-group 1035 mode active
        no shutdown

    interface Ethernet1/28
      description oradb-prd-31.office.intern interface 10 GE 1
      switchport
      switchport access vlan 850
      spanning-tree port type edge
      spanning-tree bpdufilter enable
      mtu 9214
      no shutdown

    """
    print("+------------------------------- Parsing Fex Port -----------------------+")
    primary_key = ["^interface Ethernet[0-9]+/[0-9]+/[0-9]+$"]
    secondary_keys = ["switchport$", "switchport access vlan", "description","spanning-tree port type","spanning-tree bpdufilter","switchport mode","switchport trunk allowed vlan","channel-group","shutdown","switchport trunk native vlan"]
    ethernet_port_list = {}
    parsed_output = ["config_file", "ethernet_port_id","switchport_mode", "access_vlan", "description", "spanning_tree_port_type", "bpdu_filter", "mode", "trunk_native_vlan", "trunked_vlan","port_channel_id", "shutdown_state"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_port_channel_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_port_channel_list:
            ethernet_port_id = line[0].lstrip().rstrip().split()[1].split("Ethernet")[1]
            if line[1]:
                switchport_mode = line[1][0].lstrip().rstrip()
            else:
                switchport_mode= ""
            if line[2]:
                access_vlan = line[2][0].lstrip().rstrip().split()[3]
            else:
                access_vlan = ""
            if line[3]:
                description = line[3][0].lstrip().rstrip().split("description")[1].lstrip("-").rstrip("-")
            else:
                description = ""
            if line[4]:
                spanning_tree_port_type = line[4][0].lstrip().rstrip().split("type")[1]
            else:
                spanning_tree_port_type = ""
            if line[5]:
                bpdu_filter = line[5][0].lstrip().rstrip().split()[2]
            else:
                bpdu_filter = ""
            if line[6]:
                mode = line[6][0].lstrip().rstrip().split("mode")[1]
            else:
                mode = ""
            if line[7]:
                trunked_vlan = line[7][0].lstrip().rstrip().split("vlan")[1]
                if len(line[7]) > 1:
                    for pattern in line[7][1:]:
                        trunked_add_vlan = pattern.lstrip().rstrip().split("add")[1].lstrip()
                        trunked_vlan = trunked_vlan + "," + trunked_add_vlan
            else:
                trunked_vlan = ""
            if line[8]:
                port_channel_id = line[8][0].lstrip().rstrip().split()[1]
            else:
                port_channel_id = ""
            if line[9]:
                shutdown_state = line[9][0].lstrip().rstrip()
            else:
                shutdown_state = ""
            if line[10]:
                trunk_native_vlan = line[10][0].lstrip().rstrip().split("vlan")[1].lstrip()
            else:
                trunk_native_vlan = ""
            ethernet_port_list[hash(config_file + ethernet_port_id)] = {}
            ethernet_port_list[hash(config_file + ethernet_port_id)]["ethernet_port_id"] = ethernet_port_id
            ethernet_port_list[hash(config_file + ethernet_port_id)]["switchport_mode"] = switchport_mode
            ethernet_port_list[hash(config_file + ethernet_port_id)]["access_vlan"] = access_vlan
            ethernet_port_list[hash(config_file + ethernet_port_id)]["spanning_tree_port_type"] = spanning_tree_port_type
            ethernet_port_list[hash(config_file + ethernet_port_id)]["config_file"] = config_file
            ethernet_port_list[hash(config_file + ethernet_port_id)]["description"] = description
            ethernet_port_list[hash(config_file + ethernet_port_id)]["bpdu_filter"] = bpdu_filter
            ethernet_port_list[hash(config_file + ethernet_port_id)]["mode"] = mode
            ethernet_port_list[hash(config_file + ethernet_port_id)]["trunked_vlan"] = trunked_vlan
            ethernet_port_list[hash(config_file + ethernet_port_id)]["port_channel_id"] = port_channel_id
            ethernet_port_list[hash(config_file + ethernet_port_id)]["trunk_native_vlan"] = trunk_native_vlan
            ethernet_port_list[hash(config_file + ethernet_port_id)]["shutdown_state"] = shutdown_state
    return [ethernet_port_list, parsed_output]

def parse_ethernet_port(config_files):
    """
        Parsing the following format

    interface Ethernet1/7
        description 10GE fabric to mnd-rc0002-ds173-101 - po101 - MND-RC0002-MND-RC0002-EBRR
        switchport
        switchport mode fex-fabric
        fex associate 101
        channel-group 101
        no shutdown

    interface Ethernet1/8
        description enc1-vcenet1-mgt.mnd.bss.local interface X1
        switchport
        switchport mode trunk
        switchport trunk allowed vlan 101,300,310,320,330,340,370,410,430,440,450,460,468,470,510,512,520,530,560,610,620,630,640,650,660,670,690,710,720,730,750,760,770,780,790,801-804,806-807,1014,1233,1710,1712,1720,3018
        spanning-tree port type edge trunk
        no shutdown
    interface Ethernet1/27
        description hnas2.mnd.office.intern interface 10GbE 4
        switchport
        switchport mode trunk
        switchport trunk allowed vlan 440,450,460,750
        spanning-tree port type edge
        spanning-tree bpdufilter enable
        channel-group 1035 mode active
        no shutdown

    interface Ethernet1/28
      description oradb-prd-31.office.intern interface 10 GE 1
      switchport
      switchport access vlan 850
      spanning-tree port type edge
      spanning-tree bpdufilter enable
      mtu 9214
      no shutdown

    """
    print("+--------------------------- Parsing Ethernet Port ----------------------+")
    primary_key = ["^interface Ethernet[0-9]+/[0-9]+$"]
    secondary_keys = ["switchport$", "switchport access vlan", "description","spanning-tree port type","spanning-tree bpdufilter","switchport mode","switchport trunk allowed vlan","channel-group","shutdown","switchport trunk native vlan"]
    ethernet_port_list = {}
    parsed_output = ["config_file", "ethernet_port_id", "description", "switchport_mode", "access_vlan", "spanning_tree_port_type", "bpdu_filter", "mode", "trunk_native_vlan", "trunked_vlan","port_channel_id", "shutdown_state"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_port_channel_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_port_channel_list:
            ethernet_port_id = line[0].lstrip().rstrip().split()[1].split("Ethernet")[1]
            if line[1]:
                switchport_mode = line[1][0].lstrip().rstrip()
            else:
                switchport_mode= ""
            if line[2]:
                access_vlan = line[2][0].lstrip().rstrip().split()[3]
            else:
                access_vlan = ""
            if line[3]:
                description = line[3][0].lstrip().rstrip().split("description")[1].lstrip("-").rstrip("-")
            else:
                description = ""
            if line[4]:
                spanning_tree_port_type = line[4][0].lstrip().rstrip().split("type")[1]
            else:
                spanning_tree_port_type = ""
            if line[5]:
                bpdu_filter = line[5][0].lstrip().rstrip().split()[2]
            else:
                bpdu_filter = ""
            if line[6]:
                mode = line[6][0].lstrip().rstrip().split("mode")[1]
            else:
                mode = ""
            if line[7]:
                trunked_vlan = line[7][0].lstrip().rstrip().split("vlan")[1]
                if len(line[7]) > 1:
                    for pattern in line[7][1:]:
                        trunked_add_vlan = pattern.lstrip().rstrip().split("add")[1].lstrip()
                        trunked_vlan = trunked_vlan + "," + trunked_add_vlan
            else:
                trunked_vlan = ""
            if line[8]:
                port_channel_id = line[8][0].lstrip().rstrip().split()[1]
            else:
                port_channel_id = ""
            if line[9]:
                shutdown_state = line[9][0].lstrip().rstrip()
            else:
                shutdown_state = ""
            if line[10]:
                trunk_native_vlan = line[10][0].lstrip().rstrip().split("vlan")[1].lstrip()
            else:
                trunk_native_vlan = ""
            ethernet_port_list[hash(config_file + ethernet_port_id)] = {}
            ethernet_port_list[hash(config_file + ethernet_port_id)]["ethernet_port_id"] = ethernet_port_id
            ethernet_port_list[hash(config_file + ethernet_port_id)]["switchport_mode"] = switchport_mode
            ethernet_port_list[hash(config_file + ethernet_port_id)]["access_vlan"] = access_vlan
            ethernet_port_list[hash(config_file + ethernet_port_id)]["spanning_tree_port_type"] = spanning_tree_port_type
            ethernet_port_list[hash(config_file + ethernet_port_id)]["config_file"] = config_file
            ethernet_port_list[hash(config_file + ethernet_port_id)]["description"] = description
            ethernet_port_list[hash(config_file + ethernet_port_id)]["bpdu_filter"] = bpdu_filter
            ethernet_port_list[hash(config_file + ethernet_port_id)]["mode"] = mode
            ethernet_port_list[hash(config_file + ethernet_port_id)]["trunked_vlan"] = trunked_vlan
            ethernet_port_list[hash(config_file + ethernet_port_id)]["port_channel_id"] = port_channel_id
            ethernet_port_list[hash(config_file + ethernet_port_id)]["trunk_native_vlan"] = trunk_native_vlan
            ethernet_port_list[hash(config_file + ethernet_port_id)]["shutdown_state"] = shutdown_state
    return [ethernet_port_list, parsed_output]

def parse_port_channel(config_files):
    """
        Parsing the following format

    interface port-channel1041
        description 2gb to oradb-prd-33.office.intern (VPC 1041) -  members eth101/1/36
        switchport
        switchport access vlan 790
        spanning-tree port type edge
        spanning-tree bpdufilter enable

    interface Ethernet1/10
        description enc1-vcenet1-mgt.tb.bss.local interface X3
        switchport
        switchport mode trunk
        switchport trunk allowed vlan 101,300,310,320,330,340,370,410,430,440,450,460,468,470,510,512,520,530,560,610,620,630,640,650,660,670,690,710,720,730,750,760,770,780,790,801-804,806-807,1015,1233,1710,1712,1720,3018,3901
        spanning-tree port type edge trunk
        no shutdown
    """
    print("+--------------------------- Parsing Port Channel -----------------------+")
    primary_key = ["^interface port-channel[0-9]+"]
    secondary_keys = ["switchport$", "switchport access vlan", "description","spanning-tree port type","spanning-tree bpdufilter","switchport mode","switchport trunk allowed vlan","shutdown","switchport trunk native vlan"]
    port_channel_list = {}
    parsed_output = ["config_file","port_channel_id","description", "switchport_mode", "access_vlan","mode","trunk_native_vlan","trunked_vlan","shutdown_state", "spanning_tree_port_type", "bpdu_filter"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_port_channel_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_port_channel_list:
            port_channel_id = line[0].lstrip().rstrip().split()[1].split("port-channel")[1]
            if line[1]:
                switchport_mode = line[1][0].lstrip().rstrip()
            else:
                switchport_mode= ""
            if line[2]:
                access_vlan = line[2][0].lstrip().rstrip().split()[3]
            else:
                access_vlan = ""
            if line[3]:
                description = line[3][0].lstrip().rstrip().split("description")[1].lstrip("-").rstrip("-")
            else:
                description = ""
            if line[4]:
                spanning_tree_port_type = line[4][0].lstrip().rstrip().split("type")[1]
            else:
                spanning_tree_port_type = ""
            if line[5]:
                bpdu_filter = line[5][0].lstrip().rstrip().split()[2]
            else:
                bpdu_filter = ""
            if line[6]:
                mode = line[6][0].lstrip().rstrip().split("mode")[1]
            else:
                mode = ""
            if line[7]:
                trunked_vlan = line[7][0].lstrip().rstrip().split("vlan")[1]
                if len(line[7]) > 1:
                    for pattern in line[7][1:]:
                        trunked_add_vlan = pattern.lstrip().rstrip().split("add")[1].lstrip()
                        trunked_vlan = trunked_vlan + "," + trunked_add_vlan
            else:
                trunked_vlan = ""
            if line[8]:
                shutdown_state = line[8][0]
            else:
                shutdown_state = ""
            if line[9]:
                trunk_native_vlan = line[9][0].lstrip().rstrip().split("vlan")[1].lstrip()
            else:
                trunk_native_vlan = ""
            port_channel_list[hash(config_file + port_channel_id)] = {}
            port_channel_list[hash(config_file + port_channel_id)]["port_channel_id"] = port_channel_id
            port_channel_list[hash(config_file + port_channel_id)]["switchport_mode"] = switchport_mode
            port_channel_list[hash(config_file + port_channel_id)]["access_vlan"] = access_vlan
            port_channel_list[hash(config_file + port_channel_id)]["spanning_tree_port_type"] = spanning_tree_port_type
            port_channel_list[hash(config_file + port_channel_id)]["config_file"] = [config_file]
            port_channel_list[hash(config_file + port_channel_id)]["description"] = description
            port_channel_list[hash(config_file + port_channel_id)]["bpdu_filter"] = bpdu_filter
            port_channel_list[hash(config_file + port_channel_id)]["mode"] = mode
            port_channel_list[hash(config_file + port_channel_id)]["trunked_vlan"] = trunked_vlan
            port_channel_list[hash(config_file + port_channel_id)]["shutdown_state"] = shutdown_state
            port_channel_list[hash(config_file + port_channel_id)]["trunk_native_vlan"] = trunk_native_vlan
    return [port_channel_list, parsed_output]

def parse_vlan(config_files):
    """
        Returns a dictionary of the form
    """
    print("+------------------------------- Parsing VLAN ---------------------------+")
    primary_key = ["^vlan [0-9]+"]
    secondary_keys = ["name", "mode"]
    vlan_list = {}
    parsed_output = ["config_file", "vlan_id", "name", "mode"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_vlan_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_vlan_list:
            # Expand VLAN from string like "vlan 1,2,10-14"
            if re.search(r'^vlan ([0-9]+-[0-9]+|[0-9]+,[0-9]+)', line[0], flags=0):
                vlan_id_list = split_expand_vlan_range(line[0])
                for vlan_id in vlan_id_list:
                    if hash(str(vlan_id) + config_file) not in vlan_list.keys():
                        print("Expanded Vlan %s not in the dictionnary creating new entry" %vlan_id)
                        vlan_list[hash(vlan_id + config_file)] = {}
                        vlan_list[hash(vlan_id + config_file)]["vlan_id"] = vlan_id
                        vlan_list[hash(vlan_id + config_file)]["name"] = None
                        vlan_list[hash(vlan_id + config_file)]["mode"] = None
                        vlan_list[hash(vlan_id + config_file)]["config_file"] = config_file
            else:
                vlan_id = line[0].lstrip().rstrip().split()[1]
                if line[1]:
                    name = line[1][0].lstrip().rstrip().split()[1]
                else:
                    name = ""
                if line[2]:
                    mode = line[2][0].lstrip().rstrip().split()[1]
                else:
                    mode = ""
                if hash(vlan_id+config_file) in vlan_list.keys():
                    print("Vlan %s already present in dictionnary updating info" %vlan_id)
                    if vlan_list[hash(str(vlan_id) + config_file)]["name"] is None:
                        vlan_list[hash(str(vlan_id) + config_file)]["name"] = name
                    if vlan_list[hash(str(vlan_id) + config_file)]["mode"] is None:
                        vlan_list[hash(str(vlan_id) + config_file)]["mode"] = mode
                else:
                    vlan_list[hash(vlan_id + config_file)] = {}
                    vlan_list[hash(vlan_id + config_file)]["vlan_id"] = vlan_id
                    vlan_list[hash(vlan_id + config_file)]["name"] = name
                    vlan_list[hash(vlan_id + config_file)]["mode"] = mode
                    vlan_list[hash(vlan_id + config_file)]["config_file"] = [config_file]
                    print("Vlan %s not in the dictionnary creating new entry" %vlan_id)
    return [vlan_list, parsed_output]

def parse_svi(config_files):
    """
        Returns a dictionary of the form
    """
    print("+------------------------------- Parsing SVI ----------------------------+")
    primary_key = ["^interface Vlan[0-9]+"]
    secondary_keys = ["vrf member", "ip address", "ip [0-9]+", "description", "ip access-group .+ in", "ip access-group .+ out", "ip ospf network", "ip dhcp relay address"]
    svi_list = {}
    parsed_output = ["vlan_id", "subnet", "vip", "vrf", "description", "config_file", "access_group_in", "access_group_out", "ospf_network_type", "dhcp_relay"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_svi_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_svi_list:
            vlan_id = line[0].lstrip().rstrip().split()[1].split("Vlan")[1]
            if line[1]:
                vrf = line[1][0].lstrip().rstrip().split()[2]
            else:
                vrf = ""
            if line[2]:
                interface_ip = line[2][0].lstrip().rstrip().split()[2]
                subnet = str(ipaddress.ip_interface(str(interface_ip)).network)
            else:
                subnet = ""
            if line[3]:
                vip = line[3][0].lstrip().rstrip().split()[1]
            else:
                vip = ""
            if line[4]:
                description = line[4][0].lstrip().rstrip().split("description")[1].lstrip("-").rstrip("-")
            else:
                description = ""
            if line[5]:
                access_group_in = line[5][0].lstrip().rstrip().split()[2]
            else:
                access_group_in = ""
            if line[6]:
                access_group_out = line[6][0].lstrip().rstrip().split()[2]
            else:
                access_group_out = ""
            if line[7]:
                ospf_network_type = line[7][0].lstrip().rstrip().split()[3]
            else:
                ospf_network_type = ""
            if line[8]:
                dhcp_relay_list = []
                for dhcp_relay_string in line[8]:
                    dhcp_relay_list.append(dhcp_relay_string.lstrip().rstrip().split()[4])
            else:
                dhcp_relay_list = []
            if hash(subnet+vlan_id) in svi_list.keys():
                print("Subnet %s already present in dictionnary updating site list" %subnet)
                svi_list[hash(subnet+vlan_id)]["config_file"].append(config_file)
            else:
                svi_list[hash(subnet+vlan_id)] = {}
                svi_list[hash(subnet+vlan_id)]["vlan_id"] = vlan_id
                svi_list[hash(subnet+vlan_id)]["vrf"] = vrf
                svi_list[hash(subnet+vlan_id)]["subnet"] = subnet
                svi_list[hash(subnet+vlan_id)]["vip"] = vip
                svi_list[hash(subnet+vlan_id)]["config_file"] = [config_file]
                svi_list[hash(subnet+vlan_id)]["description"] = description
                svi_list[hash(subnet+vlan_id)]["access_group_in"] = access_group_in
                svi_list[hash(subnet+vlan_id)]["access_group_out"] = access_group_out
                svi_list[hash(subnet+vlan_id)]["ospf_network_type"] = ospf_network_type
                svi_list[hash(subnet+vlan_id)]["dhcp_relay"] = dhcp_relay_list
                print("Subnet %s not in the dictionnary creating new entry" %subnet)
    return [svi_list, parsed_output]

def parse_vrf(config_files):
    """
    Parse a VRF in the form:
    vrf context v13
        description *** MPLS-VPN VRF v13 Collaboration ***
        rd 65004:13320
        address-family ipv4 unicast
            route-target import 65004:113
            route-target export 65004:113
    """
    print("+------------------------------- Parsing VRF ----------------------------+")
    primary_key = ["^vrf context"]
    secondary_keys = ["description", "rd [0-9]+", "route-target import", "route-target export"]
    vrf_list = {}
    parsed_output = ["vrf_name", "description", "route_distinguisher", "rt_import", "rt_export", "site_list"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_vrf_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_vrf_list:
            vrf_name = line[0].lstrip().rstrip().split()[2]
            if line[1]:
                description = line[1][0].split("description")[1].lstrip().rstrip()
            else:
                description = ""
            if line[2]:
                route_distinguisher = line[2][0].lstrip().rstrip().split()[1]
            else:
                route_distinguisher = ""
            if line[3]:
                rt_import = line[3][0].lstrip().rstrip().split()[2]
            else:
                rt_import = ""
            if line[4]:
                rt_export = line[4][0].lstrip().rstrip().split()[2]
            else:
                rt_export = ""
            if hash(vrf_name+route_distinguisher) in vrf_list.keys():
                print("VRF %s already present in dictionnary with same RD updating site list" %vrf_name)
                vrf_list[hash(vrf_name + route_distinguisher)]["site_list"].append(config_file)
            else:
                vrf_list[hash(vrf_name + route_distinguisher)] = {}
                vrf_list[hash(vrf_name + route_distinguisher)]["vrf_name"] = vrf_name
                vrf_list[hash(vrf_name + route_distinguisher)]["description"] = description
                vrf_list[hash(vrf_name + route_distinguisher)]["route_distinguisher"] = route_distinguisher
                vrf_list[hash(vrf_name + route_distinguisher)]["rt_import"] = rt_import
                vrf_list[hash(vrf_name + route_distinguisher)]["rt_export"] = rt_export
                vrf_list[hash(vrf_name + route_distinguisher)]["site_list"] = [config_file]
                print("VRF %s not in the dictionnary creating new entry" %vrf_name)
    return [vrf_list, parsed_output]

def parse_static_route(config_files):
    """
    Parse a static route
    vrf context v13
        ip route 10.200.192.4/32 10.200.128.5 name route2fusion_router tag 1234
    """
    print("+--------------------------- Parsing Static Route -----------------------+")
    primary_key = ["^vrf context",]
    secondary_keys = ["ip route [0-9]+"]
    static_route_list = {}
    parsed_output = ["vrf_name", "subnet", "next_hop", "route_name", "tag", "site_list"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_static_route_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_static_route_list:
            if len(line) > 1 and line[1] != None:
                for route in line[1]:
                    vrf_name = line[0].lstrip().rstrip().split()[2]
                    subnet = route.lstrip().rstrip().split()[2]
                    next_hop = route.lstrip().rstrip().split()[3]
                    if re.match(r'.+name.+', route):
                        route_name = route.split("name")[1].lstrip().rstrip().split()[0]
                    else:
                        route_name = ""
                    if re.match(r'.+tag.+', route):
                        tag = route.split("tag")[1].lstrip().rstrip().split()[0]
                    else:
                        tag = ""
                    if hash(vrf_name + subnet + next_hop) in static_route_list.keys():
                        print("Subnet %s already present in dictionnary with same VRF and next hop updating site list" %subnet)
                        static_route_list[hash(vrf_name + subnet + next_hop)]["site_list"].append(config_file)
                    else:
                        static_route_list[hash(vrf_name + subnet + next_hop)] = {}
                        static_route_list[hash(vrf_name + subnet + next_hop)]["vrf_name"] = vrf_name
                        static_route_list[hash(vrf_name + subnet + next_hop)]["subnet"] = subnet
                        static_route_list[hash(vrf_name + subnet + next_hop)]["next_hop"] = next_hop
                        static_route_list[hash(vrf_name + subnet + next_hop)]["route_name"] = route_name
                        static_route_list[hash(vrf_name + subnet + next_hop)]["tag"] = tag
                        static_route_list[hash(vrf_name + subnet + next_hop)]["site_list"] = [config_file]
                        print("Subnet %s not in the dictionnary for this VRF creating new entry" %subnet)
    return [static_route_list, parsed_output]

def parse_loopback(config_files):
    """
    Parse a Loopback interface:
    interface loopback0
        description <>
        vrf member <>
        ip address <IP>/<mask>
    """
    print("+------------------------ Parsing Loopback Interface --------------------+")
    primary_key = ["^interface loopback"]
    secondary_keys = ["description", "vrf member", "ip address"]
    loopback_list = {}
    parsed_output = ["config_file", "loopback_id", "description", "vrf", "ip_address"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_loopback_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_loopback_list:
            loopback_id = line[0].lstrip().rstrip().split()[1]
            if line[1]:
                description = line[1][0].split("description")[1].lstrip().rstrip()
            else:
                description = ""
            if line[2]:
                vrf = line[2][0].lstrip().rstrip().split()[2]
            else:
                vrf = ""
            if line[3]:
                ip_address = line[3][0].lstrip().rstrip().split()[2]
            else:
                ip_address = ""
            print("Creating loopback %s " %loopback_id)
            loopback_list[hash(config_file + loopback_id)] = {}
            loopback_list[hash(config_file + loopback_id)]["config_file"] = config_file
            loopback_list[hash(config_file + loopback_id)]["loopback_id"] = loopback_id
            loopback_list[hash(config_file + loopback_id)]["description"] = description
            loopback_list[hash(config_file + loopback_id)]["vrf"] = vrf
            loopback_list[hash(config_file + loopback_id)]["ip_address"] = ip_address
    return [loopback_list, parsed_output]

def parse_ip_access_list(config_files):
    """
    Parse an ip access-list:
    ip access-list wordpress-db
        10 permit icmp any any
        20 permit udp any eq ntp any
        30 permit udp any eq domain any
        40 permit tcp any eq domain any
        50 permit tcp any any eq 3306        
    """
    print("+-------------------------- Parsing IP Access List ----------------------+")
    primary_key = ["^ip access-list"]
    secondary_keys = ["permit", "deny", ]
    acl_list = {}
    parsed_output = ["config_file", "name", "acl_entries"]
    for config_file in config_files:
        config = read_config_file(config_file)
        print("+-- Parsing file  %s " %config_file)
        raw_acl_list = parse_config(primary_key, secondary_keys, config)
        for line in raw_acl_list:
            acl_name = line[0].lstrip().rstrip().split()[2]
            for i in range(1,len(line)):
                # Loop through each acl entry
                for acl_entry in line[i]:
                    if hash(config_file + acl_name) in acl_list.keys():
                        print("ACL %s already present in dictionary, updating acl entry list" % acl_name)
                        acl_list[hash(config_file + acl_name)]["acl_entries"].append(acl_entry.lstrip().rstrip())
                    else:
                        print("Creating ip access-list %s " %acl_name)
                        acl_list[hash(config_file + acl_name)] = {}
                        acl_list[hash(config_file + acl_name)]["config_file"] = config_file
                        acl_list[hash(config_file + acl_name)]["name"] = acl_name
                        acl_list[hash(config_file + acl_name)]["acl_entries"] = [acl_entry.lstrip().rstrip()]
    return [acl_list, parsed_output]

def parse_config(pattern_list, secondary_keys, config_file):
    """
    """
    result_list = []
    position = 0
    result_position = 0
    for line in config_file:
        for pattern in pattern_list:
            if grep(pattern, line):
                match = []
                match.append(line)
                # Initialize list with None Value for all keys
                for i in range(0, len(secondary_keys)):
                    match.append([])
                start_pattern_block = position + 1
                stop_pattern_block = start_pattern_block
                # Get start and stop indices for secondary match
                while get_indent(config_file[stop_pattern_block]) > get_indent(line):
                    stop_pattern_block = stop_pattern_block + 1
                for i in range(0, len(secondary_keys)):
                    pattern = secondary_keys[i]
                    for line in config_file[start_pattern_block:stop_pattern_block]:
                        if grep(pattern, line):
                            match[i+1].append(line)
                result_position = result_position +1
                result_list.append(match)
        position = position +1
    return result_list

def list_dir(dir_path):
    """
    """
    file_list = os.listdir(dir_path)
    return file_list

def read_arguments():
    """
    """
    parser = argparse.ArgumentParser("Usage: nxos_config_parser.py.py")
    parser.add_argument("-o", "--output-file", dest="output_file", help=" Output Excel File ", required=True)
    parser.add_argument("-d", "--conf-file-dir", dest="dir_path", help=" Dir Path ", required=True)
    args = parser.parse_args()
    return args

def main():
    """
    """
    args = read_arguments()
    orgiginal_path = os.getcwd()
    config_files = list_dir(args.dir_path)
    os.chdir(args.dir_path)
    if config_files:
        for config_file in config_files:
            config = read_config_file(config_file)
            pprint(detect_os.detect_os_type_and_version(config, config_file))
        excel_workbook = Workbook()
        vlan_dict, vlan_csv_header = parse_vlan(config_files)
        excel_workbook = create_worksheet(excel_workbook, vlan_csv_header, vlan_dict, "VLAN_audit")
        svi_dict, svi_csv_header = parse_svi(config_files)
        excel_workbook = create_worksheet(excel_workbook, svi_csv_header, svi_dict, "SVI_audit")
        vrf_dict, vrf_csv_header = parse_vrf(config_files)
        excel_workbook = create_worksheet(excel_workbook, vrf_csv_header, vrf_dict, "VRF_audit")
        static_route_dict, static_route_csv_header = parse_static_route(config_files)
        excel_workbook = create_worksheet(excel_workbook, static_route_csv_header, static_route_dict, "Static_route")
        loopback_dict, loopback_csv_header = parse_loopback(config_files)
        excel_workbook = create_worksheet(excel_workbook, loopback_csv_header, loopback_dict, "Loopback")
        port_channel_dict, port_channel_csv_header = parse_port_channel(config_files)
        excel_workbook = create_worksheet(excel_workbook, port_channel_csv_header, port_channel_dict, "port_channel")
        ethernet_port_dict, ethernet_port_csv_header = parse_ethernet_port(config_files)
        excel_workbook = create_worksheet(excel_workbook, ethernet_port_csv_header, ethernet_port_dict, "ethernet_port")
        fex_port_dict, fex_port_csv_header = parse_fex_port(config_files)
        excel_workbook = create_worksheet(excel_workbook, fex_port_csv_header, fex_port_dict, "fex_access_port")
        ip_acl_dict, ip_acl_csv_header = parse_ip_access_list(config_files)
        excel_workbook = create_worksheet(excel_workbook, ip_acl_csv_header, ip_acl_dict, "ip_access_list")
    else:
        print("Aborting Can't read input configuration file")
    os.chdir(orgiginal_path)
    excel_workbook.save(args.output_file)

if __name__ == '__main__':
    main()
