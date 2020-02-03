# Copyright (c) 2020 Cisco and/or its affiliates.
#
# This software is licensed to you under the terms of the Cisco Sample
# Code License, Version 1.1 (the "License"). You may obtain a copy of the
# License at
#
#                https://developer.cisco.com/docs/licenses
#
# All use of the material herein must be in accordance with the terms of
# the License. All rights not expressly granted by the License are
# reserved. Unless required by applicable law or agreed to separately in
# writing, software distributed under the License is distributed on an "AS
# IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
# or implied.

from openpyxl import Workbook
from openpyxl.worksheet.table import Table , TableStyleInfo
import argparse
import json
import pprint
import sys
import yaml

# Control debug
debug = False

class Logging(Exception):
    """
    Exception class handling the exception raised by this script
    """
    def fatal(self, msg):
        """
        Prints an error message and aborts program execution
        """
        sys.stderr.write(msg+"\n")
        sys.exit(1)

    def warning(self, msg):
        """
        Prints a warning message to stderr
        """
        sys.stderr.write(msg+"\n")
    
    def print_message(self, msg):
        """
        Prints a message to stdout
        """
        print(msg)

LOG = Logging()

def read_arguments():
    """
    """
    parser = argparse.ArgumentParser("Usage: state-discovery.py")
    parser.add_argument("--input", "-i", dest="input_file", help=" Discovery Input File Name", default="discovery_results.json", required=False)
    parser.add_argument("--output", "-o", dest="output_file", help=" Output Excel File Name", default="discovery_results_state.xlsx", required=False)
    parser.add_argument("--composer", "-c", dest="composer", help=" Excel Composer File", default="./composer.yaml", required=False)
    args = parser.parse_args()
    return args

def load_yaml(file):
    """
    """
    try:
        with open(file, "r") as file:
            dictionary = yaml.load(file, Loader=yaml.SafeLoader)
        file.close()
        return(dictionary)
    except:
        LOG.fatal("Error occured while importing %s as YAML input" % file)

def load_json(file):
    """
    """
    try:
        with open(file, "r") as file:
            dictionary = json.load(file)
        file.close()
        return(dictionary)
    except:
        LOG.fatal("Error occured while importing %s as JSON input" % file)

def get_child_headers(headers):
    result_headers = []
    
    # loop through entries if provided headers is a dictionary
    if type(headers) == dict:   
        for header_name in headers:
            if header_name.startswith("__"):
                child_headers = get_child_headers(headers[header_name])
                result_headers2 = result_headers + child_headers
                result_headers = result_headers2
            else:
                result_headers.append(header_name)
    else:
        LOG.warning("Non-dictionary child headers where specivied in excel composer. Entry: \"%s\"" % headers)

    return(result_headers)

def get_child_data(input_data, key_dict):
    result_entries = []


    # loop through entries, if input is a list
    if type(input_data) == list:
        
        for entry in input_data:
            result_item = {}
            
            for header_name in key_dict:
                if header_name.startswith("__"):
                    parent_header_name = header_name[2:]
                    child_results = get_child_data(entry[parent_header_name], key_dict[header_name])
                    if type(child_results) == list:
                        if len(child_results) == 1:
                            result_item.update(dict(child_results[0]))
                        else:
                            for child_entry in child_results:
                                child_entry.update(result_item)
                                result_entries.append(child_entry)
                    else:
                        ## To be implemented
                        Logging.warning("Error while getting child data in dict format")
                else:
                    try:
                        result_item[header_name] = entry[key_dict[header_name]]
                    except KeyError:
                        # Handle cases where the dictionary key are not defined
                        result_item[header_name] = ""
            result_entries.append(result_item)

    elif type(input_data) == dict:
        result_item = {}

        for header_name in key_dict:
            if header_name.startswith("__"):
                parent_header_name = header_name[2:]
                child_results = get_child_data(input_data[parent_header_name], key_dict[header_name])
                result_entries = result_entries + child_results
            else:
                try:
                    result_item[header_name] = input_data[key_dict[header_name]]
                except KeyError:
                    # Handle cases where the dictionary key are not defined
                    result_item[header_name] = ""
        result_entries.append(result_item)

    else:
        for entry in input_data:
            result_item = {}
            
            for header_name in key_dict.keys():

                if header_name.startswith("__"):
                    parent_header_name = header_name[2:]
                    child_results = get_child_data(entry[parent_header_name], entry[key_dict[header_name]])
                    result_entries = result_entries + child_results
                else:
                    try:
                        result_item[header_name] = entry[key_dict[header_name]]
                    except KeyError:
                        # Handle cases where the dictionary key are not defined
                        result_item[header_name] = ""
            result_entries.append(result_item)

    return(result_entries)

def flatten_list(baseline, input, entry_name):
    result = []
    if len(baseline) > 0:
        for b in baseline:
            for i in input:
                entry = {}
                entry = b.copy()
                if type(i) == dict:
                    entry.update(i)
                else:
                    tmp_dict = {}
                    tmp_dict[entry_name] = i
                    entry.update(tmp_dict)
                result.append(entry)
    else:
        for i in input:
            entry = {}
            if type(i) == dict:
                entry.update(i)
            else:
                tmp_dict = {}
                tmp_dict[entry_name] = i
                entry.update(tmp_dict)
            result.append(entry)

    return(result)

def analyse_data(composer, sheet_name, discovered_state):
    analysed_data = []
    # loop through discovered_state on a per-device basis
    for device in discovered_state:
        device_analysed_data = []
        LOG.print_message(" - Crunshing data from device \"%s\"" % device)
        data_entries = {}
        analyzed_entry = {}

        # Loop through the sheet keys and gather data
        for header_name in composer[sheet_name].keys():
            data_field_name = composer[sheet_name][header_name]
            
            # Handle special data field named __device_name__
            if data_field_name == "__device_name__":
                data_entries["device_name"] = device
            # Handle nested data fields (keys starting with "__")
            elif header_name.startswith("__"):
                parent_header_name = header_name[2:]
                child_data = get_child_data(discovered_state[device][parent_header_name], data_field_name)
                data_entries[header_name] = child_data
            else:
                data_entries[header_name] = discovered_state[device][data_field_name]

        # Check if any of the found data entries are lists
        list_entries = []
        non_list_entries = []
        for entry in data_entries.keys():
            if type(data_entries[entry]) == list:
                list_entries.append(entry)
            else:
                non_list_entries.append(entry)

        # Build analysed data entry
        if len(non_list_entries) > 0:       # Starting with non-list entries, as these doesn't need to be flattend
            for entry in non_list_entries:
                analyzed_entry[entry] = data_entries[entry]
            device_analysed_data.append(analyzed_entry)

        if len(list_entries) > 0:
            for entry in list_entries:          # Flatten list entries
                flattend_data = flatten_list(device_analysed_data, data_entries[entry], entry)
                device_analysed_data = flattend_data

        analysed_data = analysed_data + device_analysed_data

    return(analysed_data)

def create_worksheet(excel_workbook,sheet_name,headers,analysed_data):
    """
    """
    sheet = excel_workbook.create_sheet(title = sheet_name)
    if len(headers) > 0:
        for i in range(0,len(headers)):
            sheet.cell(column=i+1, row=1 , value = headers[i])
        row_id = 2

        for element in analysed_data:
            for i in range(0,len(headers)):
                try:
                    if len(str(element[headers[i]])) == 1:
                        sheet.cell(column = i+1, row = row_id, value = "{0}".format(str(element[headers[i]][0])))
                    else:
                        sheet.cell(column = i+1, row = row_id, value = "{0}".format(",".join(str(element[headers[i]]))))
                    sheet.cell(column = i+1, row = row_id, value = "{0}".format(str(element[headers[i]])))
                except:
                    sheet.cell(column = i+1, row = row_id, value = "")
            row_id = row_id + 1
        if len(headers) > 26:  ### Super Ugly !! Fix That
            table_cells = "A1:A" + chr(64+len(headers)-26)
        else:
            table_cells = "A1:" + chr(64+len(headers)) + str(row_id-1)
        style = TableStyleInfo(name = "TableStyleMedium9" , showRowStripes="True" )
        table = Table(displayName = sheet_name , ref = table_cells)
        table.tableStyleInfo = style
        sheet.add_table(table)
    return excel_workbook

def create_workbook(composer, discovered_state):
    excel_workbook = Workbook()
    position = 0
    for sheet in composer.keys():
        LOG.print_message("Defining headers for sheet \"%s\"" % sheet)
        # Define sheet headers
        headers = []
        for key in composer[sheet].keys():
            # Skip entries starting with "__" as this indicates that attributes are nested in the structure
            if not key.startswith("__"):
                headers.append(key)
            else:
                child_headers = get_child_headers(composer[sheet][key])
                headers = headers + child_headers

        # Analyze data
        LOG.print_message("Analyzing data for sheet \"%s\"" % sheet)
        analysed_data = analyse_data(composer, sheet, discovered_state)

        # Create worksheet
        LOG.print_message("Creating workbook for sheet \"%s\"" % sheet)
        excel_workbook = create_worksheet(excel_workbook, sheet, headers, analysed_data)

    return excel_workbook

def main():
    # Read arguments
    args = read_arguments()

    # Define dictionary to store command output
    output = {}

    # Read list of commands and inforamtion about how to compose the Excel
    LOG.print_message("Reading discovery output")
    discovery_output = load_json(args.input_file)
    LOG.print_message("Reading Excel composer")
    excel_composer = load_yaml(args.composer)

    # Compose Excel with discovered output
    excel_workbook = create_workbook(excel_composer, discovery_output)

    # Saving Excel to disk
    LOG.print_message("Saving workbook to disk")
    excel_workbook.save(args.output_file)

if __name__ == '__main__':
    main()